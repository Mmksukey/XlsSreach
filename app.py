# app.py
import os
import io
import csv
from typing import List, Dict, Any, Optional
from flask import Flask, request, render_template_string, redirect, url_for, flash, Response
from werkzeug.utils import secure_filename

# ==============================
# Config
# ==============================
UPLOAD_DIR = "uploads"
ALLOWED_EXT = {".xlsx", ".xls", ".csv", ".tsv", ".txt"}
MAX_CONTENT_MB = 50  # กันไฟล์มหึมา
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = "change-this-in-production"
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_MB * 1024 * 1024

# เก็บเฉพาะเมตาดาต้า ไม่เก็บ DataFrame ยักษ์
# item: {"filename": str, "path": str, "ext": str}
DATASTORE: List[Dict[str, Any]] = []


# ==============================
# Helpers
# ==============================
def ext_of(filename: str) -> str:
    return os.path.splitext(filename)[1].lower()

def try_decode_bytes(raw: bytes, encodings=("utf-8", "utf-8-sig", "cp874", "iso-8859-11", "cp1252")) -> str:
    for enc in encodings:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")

def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return ","


# ==============================
# Stream readers (ข้อ 4)
# ==============================
def iter_txt_hits(path: str, keyword: str, case_sensitive: bool = False):
    kw = keyword if case_sensitive else keyword.lower()
    with open(path, "rb") as f:
        raw = f.read()
    text = try_decode_bytes(raw)
    for i, line in enumerate(text.replace("\r\n", "\n").replace("\r", "\n").split("\n"), start=1):
        hay = line if case_sensitive else line.lower()
        if kw in hay:
            yield {
                "sheet": "-",
                "row": i,            # แถวข้อมูล
                "excel_row": i,      # สำหรับ TXT ให้ตรงกับบรรทัด
                "column": "text",
                "value": line
            }

def iter_csv_hits(path: str, keyword: str, case_sensitive: bool = False):
    with open(path, "rb") as f:
        raw = f.read()
    text = try_decode_bytes(raw)

    sample = text[:4096]
    delim = _sniff_delimiter(sample)

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    header = next(reader, None)
    if header is None:
        return

    col_names = [str(c) for c in header]
    kw = keyword if case_sensitive else keyword.lower()

    for r_i, row in enumerate(reader, start=2):  # excel_row ให้เริ่ม 2 เหมือนมี header
        for c_j, cell in enumerate(row):
            cell_text = "" if cell is None else str(cell)
            hay = cell_text if case_sensitive else cell_text.lower()
            if kw in hay:
                yield {
                    "sheet": "-",
                    "row": r_i - 1,
                    "excel_row": r_i,
                    "column": col_names[c_j] if c_j < len(col_names) else f"col_{c_j+1}",
                    "value": cell_text
                }

def iter_xlsx_hits(path: str, keyword: str, case_sensitive: bool = False, sheet_name: Optional[str] = None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl")

    kw = keyword if case_sensitive else keyword.lower()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        for sname in sheets:
            ws = wb[sname]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            header = [str(h) if h is not None else f"col_{i+1}" for i, h in enumerate(header)]

            for r_i, row in enumerate(rows, start=2):
                if row is None:
                    continue
                for c_j, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_text = str(cell)
                    hay = cell_text if case_sensitive else cell_text.lower()
                    if kw in hay:
                        yield {
                            "sheet": sname,
                            "row": r_i - 1,
                            "excel_row": r_i,
                            "column": header[c_j] if c_j < len(header) else f"col_{c_j+1}",
                            "value": cell_text
                        }
    finally:
        wb.close()


# ==============================
# Search core
# ==============================
def search_in_datastore(keyword: Optional[str]) -> List[Dict[str, Any]]:
    if not keyword:
        return []
    results: List[Dict[str, Any]] = []
    MAX_RESULTS = 500  # กันโตเกิน

    for item in DATASTORE:
        ext = item["ext"]
        path = item["path"]
        filename = item["filename"]

        try:
            if ext == ".txt":
                gen = iter_txt_hits(path, keyword, case_sensitive=False)
            elif ext in (".csv", ".tsv"):
                gen = iter_csv_hits(path, keyword, case_sensitive=False)
            elif ext == ".xlsx":
                gen = iter_xlsx_hits(path, keyword, case_sensitive=False)
            elif ext == ".xls":
                # ไม่สตรีมสวย แนะนำให้แปลงก่อน
                flash(f"ไฟล์ {filename} เป็น .xls แนะนำแปลงเป็น .xlsx หรือ .csv เพื่อค้นหาแบบสตรีม")
                continue
            else:
                flash(f"นามสกุลไฟล์ไม่รองรับสำหรับการค้นหาแบบสตรีม: {ext}")
                continue

            for hit in gen:
                results.append({
                    "filename": filename,
                    "sheet": hit["sheet"],
                    "column": hit["column"],
                    "value": hit["value"],
                    "data_row": hit["row"],
                    "excel_row": hit["excel_row"],
                })
                if len(results) >= MAX_RESULTS:
                    flash(f"แสดงผลสูงสุด {MAX_RESULTS} รายการ ตัดผลลัพธ์ที่เหลือ")
                    return results

        except Exception as e:
            flash(f"อ่านไฟล์ {filename} ล้มเหลว: {e}")
            continue

    return results


# ==============================
# HTML
# ==============================
INDEX_HTML = """
<!doctype html>
<html lang="th">
<head>
  <meta charset="utf-8">
  <title>ค้นหาชื่อใน Excel/CSV/TXT</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    body{font-family:system-ui,Segoe UI,Roboto,Helvetica,Arial,sans-serif; padding:24px; max-width: 1100px; margin:auto;}
    .card{border:1px solid #ddd; border-radius:12px; padding:16px; margin:16px 0;}
    table{border-collapse:collapse; width:100%;}
    th, td{border:1px solid #e5e5e5; padding:6px; text-align:left; font-size:14px;}
    th{background:#fafafa;}
    .ok{color:#0b7; font-weight:600;}
    .bad{color:#b00; font-weight:600;}
    .muted{color:#666; font-size:13px;}
    .nowrap{white-space:nowrap;}
  </style>
</head>
<body>
  <h1>ค้นหาชื่อในไฟล์ Excel/CSV/TXT</h1>

  <div class="card">
    <h3>อัปโหลดไฟล์</h3>
    <form action="{{ url_for('upload') }}" method="post" enctype="multipart/form-data">
      <input type="file" name="file" required>
      <button type="submit">อัปโหลด</button>
    </form>
    <div class="muted">รองรับ: .xlsx, .csv, .tsv, .txt (สูงสุด {{ max_mb }} MB) • .xls แนะนำให้แปลงก่อน</div>
  </div>

  <div class="card">
    <h3>ไฟล์ที่มีในระบบ</h3>
    {% if files and files|length > 0 %}
      <p>ทั้งหมด {{ files|length }} ไฟล์</p>
      <table>
        <thead>
          <tr>
            <th class="nowrap">#</th>
            <th>ไฟล์</th>
            <th class="nowrap">นามสกุล</th>
            <th>ที่เก็บ</th>
          </tr>
        </thead>
        <tbody>
          {% for f in files %}
          <tr>
            <td class="nowrap">{{ loop.index }}</td>
            <td>{{ f.filename }}</td>
            <td class="nowrap">{{ f.ext }}</td>
            <td style="max-width:480px; overflow-wrap:anywhere;">{{ f.path }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      <form action="{{ url_for('clear') }}" method="post" style="margin-top:12px">
        <button type="submit" onclick="return confirm('ล้างไฟล์ทั้งหมดใช่ไหม?')">ล้างไฟล์ทั้งหมด</button>
      </form>
    {% else %}
      <p>ยังไม่มีไฟล์ในระบบ</p>
    {% endif %}
  </div>

  <div class="card">
    <h3>ค้นหา</h3>
    <form action="{{ url_for('index') }}" method="get">
      <input type="text" name="q" placeholder="คำค้นหา" value="{{ query or '' }}">
      <button type="submit">ค้นหา</button>
    </form>
  </div>

  {% with messages = get_flashed_messages() %}
    {% if messages %}
      <div class="card">
        {% for m in messages %}
          <div>{{ m }}</div>
        {% endfor %}
      </div>
    {% endif %}
  {% endwith %}

  {% if query is not none %}
    <div class="card">
      <h3>ผลการค้นหา: "{{ query }}"</h3>
      {% if results %}
        <p class="ok">พบทั้งหมด {{ results|length }} ตำแหน่ง จาก {{ results|map(attribute='filename')|unique|list|length }} ไฟล์</p>
        <table>
          <thead>
            <tr>
              <th>ไฟล์</th>
              <th>ชีต</th>
              <th>คอลัมน์</th>
              <th>ค่า</th>
              <th>แถวข้อมูล</th>
              <th>แถว Excel</th>
            </tr>
          </thead>
          <tbody>
            {% for r in results %}
            <tr>
              <td>{{ r.filename }}</td>
              <td>{{ r.sheet }}</td>
              <td>{{ r.column }}</td>
              <td>{{ r.value }}</td>
              <td>{{ r.data_row }}</td>
              <td>{{ r.excel_row }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      {% else %}
        <p class="bad">ไม่พบผลลัพธ์</p>
      {% endif %}
    </div>
  {% endif %}
</body>
</html>
"""


# ==============================
# Routes
# ==============================
@app.get("/")
def index():
    query = request.args.get("q")
    results = search_in_datastore(query) if query else None
    return render_template_string(
        INDEX_HTML,
        query=query,
        results=results,
        files=DATASTORE,
        max_mb=MAX_CONTENT_MB
    )

@app.post("/upload")
def upload():
    f = request.files.get("file")
    if not f or f.filename == "":
        flash("ไม่พบไฟล์")
        return redirect(url_for("index"))

    filename = secure_filename(f.filename)
    ext = ext_of(filename)
    if ext not in ALLOWED_EXT:
        flash(f"ไม่รองรับไฟล์นามสกุล {ext}")
        return redirect(url_for("index"))

    save_path = os.path.join(UPLOAD_DIR, filename)
    f.save(save_path)

    # ถ้ามีชื่อไฟล์ซ้ำ ให้แทนที่เมตาดาต้าเดิม
    for i, item in enumerate(list(DATASTORE)):
        if item["filename"] == filename:
            DATASTORE.pop(i)
            break

    DATASTORE.append({
        "filename": filename,
        "path": save_path,
        "ext": ext
    })

    flash(f"อัปโหลด {filename} สำเร็จ")
    return redirect(url_for("index"))

@app.post("/clear")
def clear():
    # ลบไฟล์จริงออกจากโฟลเดอร์ด้วย
    for item in list(DATASTORE):
        try:
            if os.path.exists(item["path"]):
                os.remove(item["path"])
        except Exception:
            pass
    DATASTORE.clear()
    flash("ล้างไฟล์ทั้งหมดแล้ว")
    return redirect(url_for("index"))

# กัน 404 favicon รก log
@app.get("/favicon.ico")
def favicon():
    return Response(status=204)

if __name__ == "__main__":
    # ใช้กับ flask dev server; บน Render ให้รันผ่าน gunicorn แทน
    app.run(host="0.0.0.0", port=5000, debug=True)
